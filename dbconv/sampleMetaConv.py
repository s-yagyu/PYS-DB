"""
Read excel include meta data and convert json file

"""
from datetime import date, datetime
import json
from pathlib import Path

import openpyxl as oxl
import pandas as pd
import toml


def json_serial(obj):
    """
    exampel
    # datetime型を含むdict
    item = { "dt" : datetime.now() }

    # default引数を指定して、JSON文字列を生成
    jsonstr = json.dumps(item, default=json_serial)
    
    print(jsonstr) # '{"dt": "2017-07-05T17:01:06.112224"}'
    
    """
    if isinstance(obj, (datetime, date)):
        return obj.isoformat()
    
    else:
        return str(obj)
    
    # raise TypeError (f'Type {obj} not serializable')

def json2toml(jfile, save=False):
    """Read json file -> convert toml format

    Args:
        jfile (str or pathlib: file name
        save (bool, optional): save toml. Defaults to False.

    Returns:
        dict, toml format
        
    Example:
        JSON->Toml
        fn=r'validationData\AC3_off.json'
        jd,tm = json2toml(jfile=fn, save=True)
    """
    jpath = Path(jfile)
    with open(jpath) as file:
        # json to dict
        dict_obj = json.load(file)
        # dict to toml
        tm = toml.dumps(dict_obj)
        # print(tm)
        
        if save:
            with open(jpath.with_suffix('.toml'),'w') as f:
                toml.dump(dict_obj,f)
        
    return dict_obj, tm

def toml2json(tfile, save=False):
    """Read toml file -> convert json format

    Args:
        tfile (str or pathlib): file name
        save (bool, optional): save json. Defaults to False.

    Returns:
        dict, json
    
    Examples:
        Toml-> JSON
        fn2=r'validationData\AC3_off.toml'
        jd,js = toml2json(tfile=fn2, save=True)
    """
    tpath = Path(tfile)
    with open(tpath) as file:
        # toml to dict
        dict_obj = toml.load(file)
        # dict to toml
        js = json.dumps(dict_obj)
        # print(js)
        
        if save:
            with open(tpath.with_suffix('.json'),'w') as f:
                json.dump(dict_obj ,f, indent=4)
        
    return dict_obj, js

    
class ReadExcelSampleMeta():
    """read sample metadata excel sheet and then out metadata
    
    Input Excel data
    
    Examples:
    file_path= r"TY PYSA_v20220525_excel_invoice_0803.xlsx"
    base01=ReadExcelSampleMeta(file_path)
    base01.convert()
    
    # save each json file
    base01.each_json_out()
    
    # list meta data [dict]
    base01.data_meta_recodes
   
    # base01.length
    # print(base01.data_dict)
    # print(base01.data_dict_records[:5])
    # base01.json_out()
    # dfm = base01.df
    # print(dfm.columns)
    # print(dfm.loc[0,:])
    # orient='records'を指定すると、行ごとの辞書を保持したJSONを出力。
    # dfm.to_json('test.json',orient='records')
   
    """
  
    keys_j = ["dataLicense",
                "datasetTitle",
                "dataProvider",	
                "providerOrganization",
                "inputDate",
                "aim",
                "webReference",
                "attachedReference",
                "sampleLavel",
                "generalName",
                "sampleAbbreviation",
                "sampleDescription",
                "chemicalFormula",
                "substrateName",
                "sampleShape",
                "datFileName",
                "molFileName",
                "comment"]

    keys_e = keys_j
    
    SHEET_NAME = "記入用Sheet2"
    
    def __init__(self,filename):
        self.filename = filename
        self.file_name = Path(filename)
        
    def convert(self):
        self.sheet_name = self.find_sheetname()
        self.start_pos = self.find_keys(self.SHEET_NAME, self.keys_j[0])
        self.length = self.find_data_length(self.SHEET_NAME, self.start_pos)

        self.data_dict = self.data_meta(self.SHEET_NAME, self.start_pos , self.length)
        self.data_dict_records = self.data_meta_recordes(self.SHEET_NAME, self.start_pos , self.length)
        # self.join_meta_dict = {**self.m_meta_dict, **self.data_dict}
        # self.json = self.json_out(self.join_meta_dict)        
        self.df = self.export_df(self.data_dict)
            
    @staticmethod        
    def export_excel(meta_list, save=False, out_file_name=None):
        df_meta =  pd.DataFrame([], columns=json.loads(meta_list[0]).keys())
        for i ,dt in enumerate(meta_list):
            temp_df = pd.read_json(dt,orient='index').transpose()
            df_meta = pd.concat([df_meta, temp_df])
        
        if save:    
            if out_file_name is None:
                out_file_name = 'data_list.xlsx'

            df_meta.to_excel(out_file_name, encoding="utf-8-sig")
        
        return df_meta
    
    def export_df(self, dict_file):
        df_data =  pd.DataFrame(dict_file)
        
        return df_data
         
    def find_sheetname(self):
        # Return list of sheet name
        wb = oxl.load_workbook(self.filename, data_only=True)
        st_name_list = wb.sheetnames
        return st_name_list         
        
    def find_keys(self, sheet_name, word):
        """Find the coordinate specified by word.
        row = x, col = y
        1. read sheet specified
        2. find word in range row=1-50, col=1-20
        Excel start index -> 1
        3. return cordinate
        4. excel close

        Args:
            sheet_name (str): sheet name
            word (str): search word

        Returns:
            tuple(row, col): cordinate 
        """

        wb = oxl.load_workbook(self.filename, data_only=True)
        sheet = wb[sheet_name]
        for i in range(1,50):
            for j in range(1,20):
                get_value = sheet.cell(row=i, column=j).value
                
                if get_value == word:
                    # print(f'find word: {word}')
                    # print(f'row:{i},col:{j}')   
                    row_col =(i,j) 
                else:
                    pass
        wb.close()
        return row_col
    
    def find_data_length(self, sheet_name, start_pos, direction='row'):
        """find data length
        row direction
        col direction
        
          ------> col
         | start_pos, --, --,
         | --,
         v --, 
        row
        
        Args:
            sheet_name (str): sheet name
            start_pos (tuple): keyword codinate (row, col)

        Returns:
            int: length
        """
        wb = oxl.load_workbook(self.filename, data_only=True)
        sheet = wb[sheet_name]
        s_row, s_col = start_pos
        
        i = 0
        while True:
            
            if direction =='row':
                row=s_row+i+1
                column=s_col
            else : #direction == 'col':
                row=s_row
                column=s_col+i+1
                
            get_data = sheet.cell(row, column).value
                
            if get_data is None:
                break
            
            i = i + 1
            
        return i   
    
    def data_meta(self, sheet_name, start_pos, length):
        """ only row direction

        Args:
            sheet_name (str): sheet name
            start_pos (tuple): location
            length (int): number of record

        Returns:
            dict: 
        """
        
        s_row, s_col = start_pos
        
        wb = oxl.load_workbook(self.filename, data_only=True)
        sheet = wb[sheet_name]
        
        
        values =[]
        for k in range(len(self.keys_j)):
            value_col = []
            for j in range(length):
                get_value = sheet.cell(row=j+1+s_row, column=s_col+k).value
                value_col.append(get_value)  
            values.append(value_col)  
        
        wb.close()
            
        data_dict = dict(zip(self.keys_e, values))     
        
        return data_dict
    
    def data_meta_recordes(self, sheet_name, start_pos, length):
        """each Record
        Args:
            sheet_name (str): sheet name
            start_pos (tuple): cordinate
            length (int): number of record

        Returns:
            _type_: _description_
        """
        
        s_row, s_col = start_pos
        
        wb = oxl.load_workbook(self.filename, data_only=True)
        sheet = wb[sheet_name]
        
        
        values_record =[]
        for j in range(length):
            value_col = []
            for k in range(len(self.keys_j)):
                get_value = sheet.cell(row=j+1+s_row, column=s_col+k).value
                value_col.append(get_value)  
            values_record.append(value_col)  
        
        wb.close()
        
        data_dict_record = []
        for values in values_record:
            data_dict = dict(zip(self.keys_e, values)) 
            data_dict_record.append(data_dict)
            
        return data_dict_record
    
    def each_json_out(self): 
        
        for i, di in enumerate(self.data_dict_records):
            if 'datFileName' in self.keys_e:
                temp_file_name = di['datFileName']
                temp_file_name = f"{temp_file_name.split('.')[0]}.json"
                temp_file_name = self.file_name.with_name(temp_file_name)
            else:
                temp_file_name = f'jout_{i}.json'
            # print(temp_file_name)
            with open(temp_file_name, 'w') as f:
                json.dump(di, f, indent=4, default=json_serial)
    
    def json_out(self, dict_metadata, jsonfile_name=None):
        
        if jsonfile_name is None:
            # json_name = self.file_name.stem + sheetname + '.json'
            # jsonfile_name =self.file_name.with_name( json_name)
            jsonfile_name =self.file_name.with_suffix('.json')

        with open(jsonfile_name, 'w') as f:
            json.dump(dict_metadata, f, indent=4,default=json_serial)

        json_meta = json.dumps(dict_metadata, default=json_serial)
        
        return json_meta
    
    
# if __name__ == '__main__':
    # pass
    
    # file_path= r""
    # base01 = ReadExcelSampleMeta(file_path)
    # # print(base01.find_sheetname())

    # base01.convert()
    # # print(base01.data_dict)

    # sample_dict = base01.data_dict_records
    # print(sample_dict)
    # # save each json file
    # # base01.each_json_out()
    

    
   