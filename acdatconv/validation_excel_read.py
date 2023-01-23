# Excel data reading for verification

from datetime import date, datetime
import json
from pathlib import Path

import numpy as np
import pandas as pd
import openpyxl as oxl
import matplotlib.pyplot as plt

from acdatconv import datconv as dv



def json_serial(obj):
    """
    exampel
    # datetime型を含むdict
    item = { "dt" : datetime.now() }

    # default引数を指定して、JSON文字列を生成します
    jsonstr = json.dumps(item, default=json_serial)
    
    print(jsonstr) # '{"dt": "2017-07-05T17:01:06.112224"}'
    
    """
    if isinstance(obj, (datetime, date)):
        return obj.isoformat()
    
    else:
        return str(obj)
    
    # raise TypeError (f'Type {obj} not serializable')




class ExcelConv():
    
    keys1 =["サンプル名", "測定日時", "測定光量 [nW]", "光量補正係数名","計数時間 [sec]", "陽極電圧 [V]", "不感時間 [sec]", 
            "開始エネルギー [eV]", "終了エネルギー [eV]", "ステップ [eV]",
            "しきい値 [eV]", "傾き [Yield/eV]", "べき乗", "グラウンドレベル", "GL差分表示"]

    keys1_en =["sampleName","measureDate","uvIntensity59","nameLightCorrection","countingTime","anodeVoltage","deadTime",
               "startEnergy","finishEnergy","step",
               'thresholdEnergy','slope',"powerNumber","bg", "glDisplay"]

    key_data=['Energy[eV]', 'Yield', 'Yield^0.5']
    
    def __init__(self,filename):
        self.filename = filename
        self.file_name = Path(filename)
        
    def convert(self):
        self.sheet_name = self.find_sheetname()
        self.startp_key1 = self.find_keys(self.sheet_name[0], self.keys1[0])
        self.startp_data = self.find_keys(self.sheet_name[0], self.key_data[0])
        self.m_meta_dict = self.measure_meta(self.sheet_name[0],self.startp_key1)
        self.data_dict = self.data_meta(self.sheet_name[0], self.startp_data , self.m_meta_dict)
        self.join_meta_dict = {**self.m_meta_dict, **self.data_dict}
        self.json = self.json_out(self.join_meta_dict)
        
    def multi_sheet_convert(self):
        self.sheet_name = self.find_sheetname()
        muti_meta_list = []
        for i, fn in enumerate(self.sheet_name):
            try:
                t_startp_key1 = self.find_keys(fn, self.keys1[0])
                t_startp_data = self.find_keys(fn, self.key_data[0])
                t_m_meta_dict = self.measure_meta(fn,t_startp_key1)
                t_data_dict = self.data_meta(fn, t_startp_data , t_m_meta_dict)
                t_file_name = {'file':str(self.file_name), 'sheet':fn}
                t_join_meta_dict = {**t_file_name, **t_m_meta_dict, **t_data_dict}
                
                t_json_name = self.file_name.stem + fn + '.json'
                t_json = self.json_out(t_join_meta_dict, jsonfile_name=t_json_name)
                muti_meta_list.append(t_json)
            except:
                print(f'error sheet: {fn}')
        
        return muti_meta_list
            
    @staticmethod        
    def export_excel(meta_list, out_file_name=None):
        df_meta =  pd.DataFrame([], columns=json.loads(meta_list[0]).keys())
        for i ,dt in enumerate(meta_list):
            temp_df = pd.read_json(dt,orient='index').transpose()
            df_meta = pd.concat([df_meta, temp_df])
            
        if out_file_name is None:
            out_file_name = 'data_list.xlsx'

        df_meta.to_excel(out_file_name, encoding="utf-8-sig")
        
        return df_meta
        
    def find_sheetname(self):
        wb = oxl.load_workbook(self.filename)
        st_name_list = wb.sheetnames
        return st_name_list         
        
    def find_keys(self, sheet_name, word):
        wb = oxl.load_workbook(self.filename)
        sheet = wb[sheet_name]
        for i in range(1,50):
            for j in range(1,20):
                get_value = sheet.cell(row=i, column=j).value
                
                if get_value == word:
                    # print(f'find word: {word}')
                    # print(f'row:{i},col:{j}')   
                    row_col =(i,j) 
                else:
                    pass
        wb.close()
        return row_col
    
    def measure_meta(self, sheet_name, start_pos):
        s_row, s_col = start_pos
    
        wb = oxl.load_workbook(self.filename)
        sheet = wb[sheet_name]
        values =[]
        for i ,key in enumerate(self.keys1):
            get_value = sheet.cell(row=i+s_row, column=s_col+1).value
            values.append(get_value)
        
        wb.close()
        m_meta_dict = dict(zip(self.keys1_en, values))     
        return m_meta_dict
    
    def data_meta(self, sheet_name, start_data_pos, m_meta_dict, col_num=3):
        
        s_ene =  m_meta_dict['startEnergy']
        f_ene =  m_meta_dict['finishEnergy']
        step_ene =  m_meta_dict['step']
        
        length = int(((f_ene-s_ene) / step_ene) + 1)
        
        s_row, s_col = start_data_pos
        
        wb = oxl.load_workbook(self.filename)
        sheet = wb[sheet_name]
        
        data_title_key = []
        for i  in range(5):
            get_title = sheet.cell(row=s_row, column=s_col+i).value
            data_title_key.append(get_title)
        
        # print(data_title_key)
        
        values =[]
        for k in range(col_num):
            value_col = []
            for j in range(length):
                get_value = sheet.cell(row=j+1+s_row, column=s_col+k).value
                value_col.append(get_value)  
            values.append(value_col)  
        
        wb.close()
            
        data_dict = dict(zip(data_title_key, values))     
        
        return data_dict
    
    def json_out(self, dict_metadata, jsonfile_name=None):
        
        if jsonfile_name is None:
            # json_name = self.file_name.stem + sheetname + '.json'
            # jsonfile_name =self.file_name.with_name( json_name)
            jsonfile_name =self.file_name.with_suffix('.json')

        with open(jsonfile_name, 'w') as f:
            json.dump(dict_metadata, f, indent=4,default=json_serial)

        json_meta = json.dumps(dict_metadata, default=json_serial)
        
        return json_meta
    
if __name__ == "__main__":
    file_name =  r".\validationData\AC2S_off.xlsx"
    exdata =ExcelConv(file_name)
    # metalist=exdata.multi_sheet_convert()
    exdata.convert()
    print(exdata.m_meta_dict)
    print(exdata.data_dict)
